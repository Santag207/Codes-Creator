import qrcode
from PIL import Image
import os
import pandas as pd
from openpyxl import load_workbook

# Directorios
qr_output_dir = r"Inventario\QRs"
logo_dir = r"Inventario\Barras"

# Asegúrate de que el directorio de salida exista
os.makedirs(qr_output_dir, exist_ok=True)

# Función para generar un QR con imagen en el centro
def generate_qr_with_logo(link, document_name):
    # Ruta completa del logo y del QR
    logo_path = os.path.join(logo_dir, f"{document_name}.png")
    save_path = os.path.join(qr_output_dir, f"{document_name}.png")
    
    # Verifica si el logo existe
    if not os.path.exists(logo_path):
        print(f"Logo no encontrado: {logo_path}")
        return
    
    # Crear la instancia de QRCode
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,  # Mayor corrección de errores
        box_size=10,
        border=4,
    )
    
    # Agregar el enlace
    qr.add_data(link)
    qr.make(fit=True)

    # Generar la imagen del QR
    img = qr.make_image(fill='black', back_color='white').convert('RGB')

    # Cargar la imagen del logo y convertirla a RGBA
    logo = Image.open(logo_path).convert("RGBA")

    # Calcular el tamaño del logo (30% del tamaño del QR)
    logo_size = int(min(img.size) * 0.3)
    logo = logo.resize((logo_size, logo_size), Image.LANCZOS)

    # Calcular la posición del logo en el centro del QR
    pos = ((img.size[0] - logo_size) // 2, (img.size[1] - logo_size) // 2)

    # Crear una máscara para el logo
    logo_mask = logo.split()[3]  # Usar el canal alfa como máscara

    # Pegar el logo sobre el QR usando la máscara
    img.paste(logo, pos, mask=logo_mask)

    # Guardar la imagen
    img.save(save_path)
    print(f"QR guardado en: {save_path}")

# Cargar el archivo Excel usando openpyxl para extraer hipervínculos
file_path = r'C:\Users\santi\Desktop\Inventario PujVex.xlsx'
wb = load_workbook(file_path)
ws = wb['Inventario']  # Hoja 'Inventario'

# Extraer solo los enlaces (hipervínculos) de la columna "Links"
links_column = []

# Asumiendo que los links están en la columna 'Links' (ajusta min_col y max_col según la columna correcta)
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=15, max_col=15):  # Ajusta la columna si es necesario
    for cell in row:
        if cell.hyperlink:
            links_column.append(cell.hyperlink.target)  # Extrae el link completo
        else:
            links_column.append("")  # Si no hay link, deja vacío

# Usar la columna 'Codigo Pieza' de tu archivo Excel (sigue usando pandas para esto)
df = pd.read_excel(file_path, sheet_name='Inventario')
document_names = df['Codigo Pieza']  # Columna con los nombres de las piezas

# Verificar que el número de links y de nombres sea el mismo
if len(links_column) != len(document_names):
    print(f"Error: La cantidad de enlaces ({len(links_column)}) y nombres ({len(document_names)}) no coincide.")
else:
    # Imprimir para verificar
    print("Verificación de enlaces y nombres:")
    for link, doc_name in zip(links_column, document_names):
        print(f"Enlace: {link}, Nombre de la pieza: {doc_name}")
    
    # Iterar sobre los datos y generar los QR automáticamente
    for link, doc_name in zip(links_column, document_names):
        if link and doc_name:  # Solo generar si hay un enlace y un nombre válidos
            generate_qr_with_logo(link, doc_name)
